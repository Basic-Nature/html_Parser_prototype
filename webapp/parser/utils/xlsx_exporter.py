from __future__ import annotations

from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import re
BASE_COLUMNS = {"Precinct", "Total Ballots Reported", "Percent Reported"}

def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = cell.value
            if val is None:
                continue
            ln = len(str(val))
            if ln > max_len:
                max_len = ln
        ws.column_dimensions[col_letter].width = min( max_len + 2, 55)

def _apply_styles(ws, header_rows: int):
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    fill_top = PatternFill("solid", fgColor="F2F2F7")
    fill_sub = PatternFill("solid", fgColor="FAFAFD")

    for r in range(1, header_rows + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = header_font
            cell.alignment = center
            cell.border = border
            cell.fill = fill_top if r == 1 else fill_sub

    # Body
    for r in range(header_rows + 1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")

def export_candidate_group_pivot_xlsx(
    *,
    flat_headers: List[str],
    rows: List[Dict[str, Any]],
    hierarchical_header_rows: Optional[List[List[str]]] = None,
    xlsx_path: str,
    context: Optional[dict] = None,
    format_numbers: bool = True,
    apply_color_scale: bool = True
) -> str:
    """
    Export table with optional two-row hierarchical headers.
    If hierarchical_header_rows is None, builds a simple single-row header sheet.
    """
    context = context or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    if hierarchical_header_rows and len(hierarchical_header_rows) >= 2:
        row1, row2 = hierarchical_header_rows[0], hierarchical_header_rows[1]
        # Safety alignment
        if len(row1) != len(flat_headers):
            row1 = flat_headers
        if len(row2) != len(flat_headers):
            row2 = [""] * len(flat_headers)

        # Write two header rows
        ws.append(row1)
        ws.append(row2)

        # Merge contiguous candidate blocks in first row where row2 has values
        current_label = None
        start_col = None
        for idx, (h1, h2) in enumerate(zip(row1, row2), start=1):
            label = h1
            # For base columns (blank second row) do not merge across
            if h1 in BASE_COLUMNS and (h2 == "" or h2 is None):
                if start_col and current_label:
                    if start_col != idx - 1:
                        ws.merge_cells(start_row=1, start_column=start_col,
                                       end_row=1, end_column=idx - 1)
                current_label = None
                start_col = None
                continue
            if current_label is None:
                current_label = label
                start_col = idx
            elif label != current_label:
                # close previous
                if start_col is not None and start_col != idx - 1:
                    ws.merge_cells(start_row=1, start_column=start_col,
                                   end_row=1, end_column=idx - 1)
                current_label = label
                start_col = idx
        # Close tail
        if current_label and start_col and start_col != len(flat_headers):
            ws.merge_cells(start_row=1, start_column=start_col,
                           end_row=1, end_column=len(flat_headers))
        header_rows_count = 2
        data_start_row = 3
    else:
        # Simple header
        ws.append(flat_headers)
        header_rows_count = 1
        data_start_row = 2

    # Write data rows (raw first)
    numeric_candidate_re = re.compile(r"(Total Reported|Total Vote|Grand Total|Cum Vote)$", re.I)
    percent_col_re = re.compile(r"(Percent Reported|% Vote|Cum %)$", re.I)
    for r in rows:
        ws.append([r.get(h, "") for h in flat_headers])

    if format_numbers:
        # Determine column types
        for col_idx, header in enumerate(flat_headers, start=1):
            col_letter = get_column_letter(col_idx)
            is_percent = bool(percent_col_re.search(header))
            is_numeric = bool(numeric_candidate_re.search(header))
            # Fallback heuristic: if >70% numeric strings
            if not (is_percent or is_numeric):
                numeric_hits = 0
                non_empty = 0
                for row_idx in range(1 + (2 if hierarchical_header_rows else 1), ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    val = cell.value
                    if val not in (None, ""):
                        non_empty += 1
                        if isinstance(val, (int, float)) or (isinstance(val, str) and re.fullmatch(r"\d{1,3}(,\d{3})*|\d+(\.\d+)?", val.strip())):
                            numeric_hits += 1
                if non_empty and numeric_hits / non_empty >= 0.7:
                    is_numeric = True
            # Convert & format
            for row_idx in range(1 + (2 if hierarchical_header_rows else 1), ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value
                if val in (None, ""):
                    continue
                if is_percent:
                    # Accept "85", "85%", "85.2", "85.2%"
                    if isinstance(val, str):
                        sv = val.strip().replace("%", "")
                        sv = sv.replace(",", "")
                        try:
                            f = float(sv) / 100.0
                            cell.value = f
                            cell.number_format = "0.0%"
                        except Exception:
                            pass
                    elif isinstance(val, (int, float)):
                        cell.value = float(val) / (100.0 if val > 1 else 1.0)
                        cell.number_format = "0.0%"
                elif is_numeric:
                    if isinstance(val, str):
                        sv = val.replace(",", "").strip()
                        if re.fullmatch(r"-?\d+", sv):
                            try:
                                cell.value = int(sv)
                                cell.number_format = "#,##0"
                            except Exception:
                                pass
                        elif re.fullmatch(r"-?\d+\.\d+", sv):
                            try:
                                cell.value = float(sv)
                                cell.number_format = "#,##0.00"
                            except Exception:
                                pass
                    elif isinstance(val, (int, float)):
                        cell.number_format = "#,##0" if isinstance(val, int) else "#,##0.00"

    if apply_color_scale:
        data_row_start = 3 if (hierarchical_header_rows and len(hierarchical_header_rows) >= 2) else 2
        data_row_end = ws.max_row
        if data_row_end >= data_row_start:
            # Apply separate scales to counts and percents
            count_fill_cols = []
            percent_fill_cols = []
            for col_idx, header in enumerate(flat_headers, start=1):
                if percent_col_re.search(header):
                    percent_fill_cols.append(col_idx)
                elif numeric_candidate_re.search(header):
                    count_fill_cols.append(col_idx)
            # Counts scale (greens)
            for ci in count_fill_cols:
                rng = f"{get_column_letter(ci)}{data_row_start}:{get_column_letter(ci)}{data_row_end}"
                ws.conditional_formatting.add(
                    rng,
                    ColorScaleRule(
                        start_type="min", start_color="F9F9F9",
                        mid_type="percentile", mid_value=50, mid_color="B7E1CD",
                        end_type="max", end_color="228B22"
                    )
                )
            # Percent scale (blue)
            for ci in percent_fill_cols:
                rng = f"{get_column_letter(ci)}{data_row_start}:{get_column_letter(ci)}{data_row_end}"
                ws.conditional_formatting.add(
                    rng,
                    ColorScaleRule(
                        start_type="min", start_color="F9F9FF",
                        mid_type="percentile", mid_value=50, mid_color="A8C8FF",
                        end_type="max", end_color="004A99"
                    )
                )

    _apply_styles(ws, header_rows=header_rows_count)
    _auto_width(ws)

    # Freeze below headers
    ws.freeze_panes = ws.cell(row=data_start_row, column=1)

    # Add metadata sheet (light)
    meta_ws = wb.create_sheet("Meta")
    meta_ws.append(["Key", "Value"])
    meta_items = {
        "pivot_mode": "candidate_groups_rawjson" if "candidate_groups_rawjson" in (context.get("pivot_modes") or []) else "",
        "contest": context.get("contest"),
        "structure_hash": context.get("structure_hash"),
        "reporting_percent": context.get("summary", {}).get("contest_reporting_percent"),
        "candidate_count": context.get("summary", {}).get("candidate_count"),
    }
    for k, v in meta_items.items():
        meta_ws.append([k, v])

    wb.save(xlsx_path)
    return xlsx_path